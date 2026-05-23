"""Excel questionnaire workbook for non-scriptable Mythos controls.

Reuses the styling vocabulary from ``excel_reporter`` and follows the MRRA
Excel template conventions (per-section sheets, blank answer cells for
workshop completion, scanner findings pre-filled).
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from znextscan.reporters.excel_reporter import (
    HEADER_FILL,
    HEADER_FONT,
    P0_FILL,
    P1_FILL,
    STATUS_FILLS,
    STATUS_FONTS,
    THIN_BORDER,
)

_DIM_NAMES = {
    "R": "R - Preparedness & Readiness",
    "C": "C - Controls to Add",
    "V": "V - Vulnerability Patching",
    "X": "X - Response & Recovery",
}

_COLUMNS = [
    ("Control ID", 12),
    ("Name", 42),
    ("Pri", 5),
    ("NIST", 10),
    ("Validation", 11),
    ("Question", 60),
    ("Evidence Prompt", 34),
    ("Prefilled", 10),
    ("Answer", 16),
    ("Notes", 30),
]


def write_questionnaire_excel(questionnaire: dict[str, Any], path: str | Path) -> Path:
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Instructions"
    ws0.column_dimensions["A"].width = 100
    ws0["A1"] = "zNextScan — Mythos Assessment Questionnaire"
    ws0["A1"].font = Font(size=15, bold=True, color="FF6B35")
    ws0["A3"] = (
        "Complete one sheet per dimension during the workshop. Scanner/Hybrid "
        "controls that produced an automated result show a 'Prefilled' status; "
        "Interview/Document controls have blank Answer cells for you to fill."
    )
    ws0["A3"].alignment = Alignment(wrap_text=True, vertical="top")
    ws0["A5"] = (
        f"Assessment: {questionnaire['assessment']}    "
        f"Generated: {questionnaire['generated_at']}"
    )

    for section in questionnaire["sections"]:
        dim = section["dimension"]
        ws = wb.create_sheet(_DIM_NAMES.get(dim, dim))
        ws["A1"] = _DIM_NAMES.get(dim, dim)
        ws["A1"].font = Font(size=13, bold=True, color="FF6B35")

        header_row = 3
        for col, (title, width) in enumerate(_COLUMNS, 1):
            cell = ws.cell(row=header_row, column=col, value=title)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="left")
            ws.column_dimensions[get_column_letter(col)].width = width

        row = header_row + 1
        for c in section["controls"]:
            values = [
                c["control_id"],
                c["name"],
                c["priority"],
                c["nist_function"],
                c["validation_method"],
                c["question"],
                c["evidence_prompt"],
                c["prefilled_status"] or "",
                "",
                "",
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = THIN_BORDER
            p_cell = ws.cell(row=row, column=3)
            if c["priority"] == "P0":
                p_cell.fill = P0_FILL
            elif c["priority"] == "P1":
                p_cell.fill = P1_FILL
            status = c["prefilled_status"]
            if status and status in STATUS_FILLS:
                status_cell = ws.cell(row=row, column=8)
                status_cell.fill = STATUS_FILLS[status]
                status_cell.font = STATUS_FONTS[status]
            row += 1

    p = Path(path)
    wb.save(p)
    return p
