"""Excel report generation — full 50-control MRRA checklist with scanner results.

Generates an Excel workbook with:
- Summary sheet with readiness score
- Per-NIST-function sheets (F1-F10) with all 50 controls
- Scanner-automated controls are pre-filled with results
- Interview/document controls are left blank for manual assessment
"""

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from znextscan.scanner import ScanResult

# All 50 MRRA controls organized by NIST function
# (control_id, name, priority, validation_method, stride)
MRRA_CONTROLS: list[tuple[str, str, list[tuple[str, str, str, str, str]]]] = [
    (
        "F1 - Asset Inventory",
        "IDENTIFY",
        [
            ("ID-001", "LPAR Inventory with Criticality Ratings", "P0", "Interview + Doc", "All"),
            (
                "ID-002",
                "Privileged User Inventory (SPECIAL/OPS/AUDITOR)",
                "P0",
                "Scanner",
                "Spoofing, Priv Esc",
            ),
            ("ID-003", "APF Library Inventory", "P0", "Scanner", "Priv Esc"),
            ("ID-004", "Network Entry Points Mapping", "P1", "Interview + Doc", "All"),
            (
                "ID-005",
                "Data Classification (PCI, PHI, PII)",
                "P0",
                "Interview + Doc",
                "Info Disclosure",
            ),
            ("ID-006", "Critical Business Service Mapping", "P1", "Interview", "All"),
        ],
    ),
    (
        "F2 - Identity & Access",
        "PROTECT",
        [
            ("IAM-001", "Multi-Factor Authentication Coverage", "P0", "Interview", "Spoofing"),
            ("IAM-002", "Strong Password Policy (SETROPTS)", "P0", "Scanner", "Spoofing"),
            ("IAM-003", "Default Account Removal (IBMUSER, SYS1)", "P0", "Scanner", "Spoofing"),
            ("IAM-004", "RACF SPECIAL Attribute Restriction (<5)", "P0", "Scanner", "Priv Esc"),
            ("IAM-005", "Started Task (STC) Security", "P1", "Scanner", "Priv Esc"),
            ("IAM-006", "Console Security (CN00 Restriction)", "P1", "Interview", "Priv Esc"),
            ("IAM-007", "Shared Userid Elimination", "P0", "Interview", "Repudiation"),
            ("IAM-008", "Batch Job RACF Enforcement (BATCHALLRACF)", "P1", "Scanner", "Priv Esc"),
        ],
    ),
    (
        "F3 - Encryption",
        "PROTECT",
        [
            ("ENC-001", "Pervasive Encryption Enabled", "P1", "Interview", "Info Disclosure"),
            ("ENC-002", "ICSF Key Management", "P0", "Scanner", "Info Disclosure"),
            ("ENC-003", "Critical Dataset Encryption", "P1", "Interview", "Info Disclosure"),
            ("ENC-004", "Network Encryption (TN3270E, SFTP)", "P0", "Interview", "Info Disclosure"),
            ("ENC-005", "Crypto Hardware Utilization", "P1", "Scanner", "Info Disclosure"),
        ],
    ),
    (
        "F4 - Network",
        "PROTECT",
        [
            ("NET-001", "LPAR Isolation (Prod/Dev/Test)", "P0", "Interview + Doc", "All"),
            ("NET-002", "Network Segmentation (TCP/IP)", "P1", "Interview", "All"),
            ("NET-003", "External Firewall Rules", "P0", "Interview + Doc", "All"),
            ("NET-004", "Administrative Jump Server Access", "P1", "Interview", "Spoofing"),
        ],
    ),
    (
        "F5 - Backup & Recovery",
        "PROTECT & RECOVER",
        [
            (
                "BCK-001",
                "Immutable Backups (Safeguarded Copy/LWORM)",
                "P0",
                "Interview",
                "DoS, Tampering",
            ),
            ("BCK-002", "Backup Air-Gap (Network Isolation)", "P0", "Interview", "DoS"),
            (
                "BCK-003",
                "Critical System Backup (RACF DB, Catalogs)",
                "P0",
                "Interview",
                "DoS, Priv Esc",
            ),
            ("BCK-004", "RTO/RPO Definition per Workload", "P1", "Interview + Doc", "DoS"),
            ("BCK-005", "DR Recovery Testing (Last 12 months)", "P0", "Interview + Doc", "DoS"),
            ("BCK-006", "Physical/Offline Tape Backups", "P1", "Interview", "DoS"),
        ],
    ),
    (
        "F6 - Patch Management",
        "PROTECT",
        [
            ("PTH-001", "z/OS Patch Management Process (PTF/RSU)", "P0", "Interview", "All"),
            ("PTH-002", "USS & Middleware Patching", "P0", "Interview", "All"),
            ("PTH-003", "Vulnerability Scanning", "P1", "Interview", "All"),
            ("PTH-004", "CVE Tracking Process", "P1", "Interview", "All"),
        ],
    ),
    (
        "F7 - Monitoring & Detection",
        "DETECT",
        [
            (
                "MON-001",
                "SMF Recording (Critical Types 80, 83, 30)",
                "P0",
                "Scanner",
                "Repudiation",
            ),
            ("MON-002", "SMF Forwarding to SIEM", "P0", "Interview", "All"),
            ("MON-003", "RACF Audit Logging (SETROPTS AUDIT)", "P0", "Scanner", "Repudiation"),
            ("MON-004", "Failed Logon Alerting", "P1", "Interview", "Spoofing"),
            ("MON-005", "Privileged Account Monitoring", "P1", "Interview", "Priv Esc"),
            ("MON-006", "Console Activity Logging", "P1", "Interview", "Repudiation"),
            ("MON-007", "24/7 SOC Coverage for Mainframe", "P0", "Interview", "All"),
        ],
    ),
    (
        "F8 - Incident Response",
        "RESPOND",
        [
            ("IR-001", "Mainframe-Specific IR Playbook", "P0", "Document", "DoS"),
            ("IR-002", "Emergency Break-Glass Credentials", "P0", "Interview", "DoS"),
            ("IR-003", "Network/LPAR Isolation Procedures", "P0", "Document", "DoS"),
            ("IR-004", "Recovery Procedures (RACF, Datasets)", "P0", "Document", "DoS"),
            ("IR-005", "Tabletop Exercise (Last 12 months)", "P1", "Interview + Doc", "All"),
        ],
    ),
    (
        "F9 - Training",
        "PROTECT",
        [
            ("TRN-001", "Mainframe Security Training (Sysprogs/DBAs)", "P1", "Interview", "All"),
            ("TRN-002", "RACF Admin Security Training", "P0", "Interview", "Priv Esc"),
            ("TRN-003", "Security Awareness (Phishing for Admins)", "P1", "Interview", "Spoofing"),
        ],
    ),
    (
        "F10 - Third-Party",
        "IDENTIFY & PROTECT",
        [
            (
                "3PT-001",
                "MSP/Vendor Access Controls (MFA + Time-Limited)",
                "P0",
                "Interview",
                "Spoofing, Priv Esc",
            ),
            ("3PT-002", "ISV Product Patching", "P1", "Interview", "All"),
            ("3PT-003", "Remote Support Access Logging", "P1", "Interview", "Repudiation"),
        ],
    ),
    (
        "F11 - System Integrity",
        "PROTECT",
        [
            ("SCI-001", "APF Library Integrity Monitoring", "P0", "Scanner", "Tampering, Priv Esc"),
            ("SCI-002", "System Library Change Control", "P0", "Interview + Doc", "Tampering"),
            ("SCI-003", "File Integrity Monitoring (FIM)", "P1", "Interview", "Tampering"),
            ("SCI-004", "Program Control (RACF PROGRAM Class)", "P1", "Scanner", "Priv Esc"),
            (
                "SCI-005",
                "RACF Database Protection (UACC=NONE)",
                "P0",
                "Scanner",
                "Tampering, Priv Esc",
            ),
            ("SCI-006", "ICHAUTAB Authorization Bypass Table", "P1", "Scanner", "Priv Esc"),
        ],
    ),
]

# Map MRRA control IDs to scanner check IDs where they differ
# (scanner uses EXT-xxx internally, MRRA framework uses domain-specific IDs)
MRRA_TO_SCANNER_ID: dict[str, str] = {
    "IAM-008": "EXT-023",  # Batch RACF Enforcement
    "SCI-005": "EXT-024",  # RACF Database Protection
    "SCI-006": "EXT-025",  # ICHAUTAB Authorization Bypass
}

# Styling
STATUS_FILLS = {
    "Pass": PatternFill(start_color="C6EFCE", fill_type="solid"),
    "Partial": PatternFill(start_color="FFEB9C", fill_type="solid"),
    "Fail": PatternFill(start_color="FFC7CE", fill_type="solid"),
    "Error": PatternFill(start_color="E2BFFF", fill_type="solid"),
    "Skipped": PatternFill(start_color="F2F2F2", fill_type="solid"),
}
STATUS_FONTS = {
    "Pass": Font(color="006100"),
    "Partial": Font(color="9C6500"),
    "Fail": Font(color="9C0006"),
    "Error": Font(color="7030A0"),
    "Skipped": Font(color="7F7F7F"),
}
P0_FILL = PatternFill(start_color="FFE6E6", fill_type="solid")
P1_FILL = PatternFill(start_color="FFF9E6", fill_type="solid")
HEADER_FILL = PatternFill(start_color="1F1F1F", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
ORANGE_FONT = Font(color="FF6B35", bold=True)
SCANNER_FILL = PatternFill(start_color="F0F7FF", fill_type="solid")
THIN_BORDER = Border(bottom=Side(style="thin", color="E5E7EB"))


def write_excel_report(scan: ScanResult, output_path: str | Path) -> Path:
    """Generate full 50-control MRRA Excel workbook with scanner results."""
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    d = scan.to_dict()
    meta = d["scan_metadata"]
    summary = d["summary"]

    # Build lookup of scanner results by control ID
    scanner_results: dict[str, dict] = {c["control_id"]: c for c in d["controls"]}

    wb = Workbook()

    # Summary sheet
    ws = wb.active
    assert ws is not None  # a new workbook always has an active sheet
    ws.title = "Summary"
    _build_summary(ws, meta, summary)

    # Per-function sheets with all 50 controls
    for func_name, nist, controls in MRRA_CONTROLS:
        sheet_name = func_name[:31]
        ws_func = wb.create_sheet(sheet_name)
        _build_function_sheet(ws_func, func_name, nist, controls, scanner_results)

    # Extended scanner checks (not in the 50 standard controls)
    # EXT checks that are NOT mapped to MRRA standard controls go in Extended tab
    mapped_ext_ids = set(MRRA_TO_SCANNER_ID.values())
    ext_checks = [
        c
        for c in d["controls"]
        if c["control_id"].startswith("EXT-") and c["control_id"] not in mapped_ext_ids
    ]
    if ext_checks:
        ws_ext = wb.create_sheet("Extended Checks")
        _build_extended_sheet(ws_ext, ext_checks)

    wb.save(str(path))
    return path


def _build_summary(ws: Any, meta: dict, summary: dict) -> None:
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 40

    ws.merge_cells("A1:B1")
    ws["A1"] = "MRRA — Mainframe Ransomware Readiness Assessment"
    ws["A1"].font = Font(size=16, bold=True, color="FF6B35")

    rows = [
        ("", ""),
        ("Scan ID", meta["scan_id"][:8]),
        ("Date", meta["start_time"][:10]),
        ("Connection", meta["connection_method"].upper()),
        ("Duration", f"{meta.get('duration_seconds', 0):.1f}s"),
        ("Scanner Version", meta["scanner_version"]),
        ("", ""),
        ("READINESS SCORE", f"{summary['score_pct']}%"),
        ("", ""),
        ("Scanner Checks", ""),
        ("  Passed", summary["passed"]),
        ("  Partial", summary["partial"]),
        ("  Failed", summary["failed"]),
        ("  Errors", summary["errors"]),
        ("  Skipped", summary["skipped"]),
        ("  Total Automated", summary["total_checks"]),
        ("", ""),
        ("Manual Checks", "23 controls — complete during workshop"),
        ("Total MRRA Controls", "50"),
    ]

    for i, (label, value) in enumerate(rows, 2):
        ws[f"A{i}"] = label
        ws[f"B{i}"] = value
        if label == "READINESS SCORE":
            ws[f"A{i}"].font = Font(bold=True, size=14)
            ws[f"B{i}"].font = Font(bold=True, size=14, color="FF6B35")
        elif label.startswith("  "):
            ws[f"A{i}"].font = Font(color="555555")


def _build_function_sheet(
    ws: Any,
    func_name: str,
    nist: str,
    controls: list[tuple[str, str, str, str, str]],
    scanner_results: dict[str, dict],
) -> None:
    headers = [
        ("Control ID", 13),
        ("Control Name", 42),
        ("Priority", 9),
        ("Validation", 14),
        ("STRIDE", 18),
        ("Status", 10),
        ("Finding Notes", 55),
        ("Evidence", 15),
    ]

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = f"{func_name}  ({nist})"
    ws["A1"].font = Font(size=13, bold=True, color="FF6B35")

    auto_count = sum(1 for cid, *_ in controls if cid in scanner_results)
    manual_count = len(controls) - auto_count
    ws["A2"] = (
        f"{len(controls)} controls: {auto_count} scanner-automated, {manual_count} manual (interview/document)"
    )
    ws["A2"].font = Font(color="9CA3AF", size=10)

    # Headers at row 4
    for col, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left")
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A5"

    for row_idx, (cid, name, priority, validation, stride) in enumerate(controls, 5):
        # Control ID
        ws.cell(row=row_idx, column=1, value=cid).font = ORANGE_FONT
        # Name
        ws.cell(row=row_idx, column=2, value=name)
        # Priority with color
        p_cell = ws.cell(row=row_idx, column=3, value=priority)
        if priority == "P0":
            p_cell.fill = P0_FILL
            p_cell.font = Font(bold=True, color="9C0006")
        elif priority == "P1":
            p_cell.fill = P1_FILL
            p_cell.font = Font(bold=True, color="9C6500")
        # Validation method
        ws.cell(row=row_idx, column=4, value=validation)
        # STRIDE
        ws.cell(row=row_idx, column=5, value=stride).font = Font(color="9CA3AF", size=10)

        # Scanner result — check direct ID or mapped EXT ID
        scanner_id = MRRA_TO_SCANNER_ID.get(cid, cid)
        result = scanner_results.get(cid) or scanner_results.get(scanner_id)
        if result:
            status = result["status"]
            s_cell = ws.cell(row=row_idx, column=6, value=status)
            s_cell.fill = STATUS_FILLS.get(status, PatternFill())
            s_cell.font = STATUS_FONTS.get(status, Font())
            findings = "\n".join(result.get("findings", []))
            ws.cell(row=row_idx, column=7, value=findings).alignment = Alignment(
                wrap_text=True, vertical="top"
            )
            # Light blue background for scanner-filled rows
            for c in range(1, 9):
                if c not in (3, 6):  # don't override priority/status colors
                    cell = ws.cell(row=row_idx, column=c)
                    if not cell.fill or cell.fill.start_color.index == "00000000":
                        cell.fill = SCANNER_FILL
        else:
            # Manual control — leave status blank for assessor
            ws.cell(row=row_idx, column=6, value="").alignment = Alignment(horizontal="center")
            ws.cell(row=row_idx, column=7, value="")

        ws.cell(row=row_idx, column=8, value="")

        for col in range(1, 9):
            ws.cell(row=row_idx, column=col).border = THIN_BORDER


def _build_extended_sheet(ws: Any, checks: list[dict]) -> None:
    """Sheet for extended scanner checks (EXT-001 to EXT-015)."""
    ws.merge_cells("A1:E1")
    ws["A1"] = "Extended Scanner Checks (beyond standard 50 controls)"
    ws["A1"].font = Font(size=13, bold=True, color="FF6B35")

    headers = [("Control ID", 13), ("Name", 35), ("Status", 10), ("Findings", 60)]
    for col, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A4"

    for row, check in enumerate(checks, 4):
        status = check["status"]
        ws.cell(row=row, column=1, value=check["control_id"]).font = ORANGE_FONT
        ws.cell(row=row, column=2, value=check["control_name"])
        s_cell = ws.cell(row=row, column=3, value=status)
        s_cell.fill = STATUS_FILLS.get(status, PatternFill())
        s_cell.font = STATUS_FONTS.get(status, Font())
        ws.cell(row=row, column=4, value="\n".join(check.get("findings", []))).alignment = (
            Alignment(wrap_text=True, vertical="top")
        )
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = THIN_BORDER
